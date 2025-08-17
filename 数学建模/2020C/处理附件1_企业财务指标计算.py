#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
附件1企业财务指标计算程序
根据发票数据计算主营业务收入、毛利率、信誉评级、发票作废率
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

def load_data():
    """加载附件1数据"""
    try:
        # 读取附件1数据
        df = pd.read_excel('附件1：123家有信贷记录企业的相关数据.xlsx')
        print(f"成功加载数据，共{len(df)}家企业")
        print(f"数据列名：{list(df.columns)}")
        return df
    except Exception as e:
        print(f"数据加载失败：{e}")
        return None

def calculate_business_indicators(df):
    """
    根据附件公式计算企业财务指标
    公式：
    1. 主营业务收入 = 销项发票金额之和
    2. 毛利率 = (主营业务收入 - 主营业务成本) / 主营业务收入 * 100%
    3. 作废发票率 = 总作废发票数量 / 总发票数量 * 100%
    """
    
    # 初始化结果DataFrame
    results = pd.DataFrame()
    results['企业编号'] = df['企业编号'] if '企业编号' in df.columns else range(1, len(df)+1)
    
    # 1. 计算主营业务收入（销项发票金额之和）
    # 查找销项发票相关列
    sales_invoice_cols = [col for col in df.columns if '销项' in col and ('金额' in col or '收入' in col)]
    if sales_invoice_cols:
        results['主营业务收入'] = df[sales_invoice_cols].sum(axis=1)
    else:
        # 如果没有明确的销项发票列，尝试查找收入相关列
        revenue_cols = [col for col in df.columns if any(keyword in col for keyword in ['收入', '营业额', '销售额'])]
        if revenue_cols:
            results['主营业务收入'] = df[revenue_cols[0]]
        else:
            results['主营业务收入'] = 0
            print("警告：未找到销项发票金额相关数据")
    
    # 2. 计算毛利率
    # 查找成本相关列
    cost_cols = [col for col in df.columns if any(keyword in col for keyword in ['成本', '进项'])]
    if cost_cols:
        # 主营业务成本
        results['主营业务成本'] = df[cost_cols[0]]
        # 计算毛利率
        results['毛利率'] = np.where(
            results['主营业务收入'] > 0,
            (results['主营业务收入'] - results['主营业务成本']) / results['主营业务收入'] * 100,
            0
        )
    else:
        # 如果没有成本数据，假设毛利率为15%（行业平均水平）
        results['主营业务成本'] = results['主营业务收入'] * 0.85
        results['毛利率'] = 15.0
        print("警告：未找到成本数据，使用估算毛利率15%")
    
    # 3. 计算发票作废率
    # 查找作废发票相关列
    void_invoice_cols = [col for col in df.columns if '作废' in col]
    total_invoice_cols = [col for col in df.columns if any(keyword in col for keyword in ['总发票', '发票总数', '发票数量'])]
    
    if void_invoice_cols and total_invoice_cols:
        results['作废发票数量'] = df[void_invoice_cols[0]]
        results['总发票数量'] = df[total_invoice_cols[0]]
        results['发票作废率'] = np.where(
            results['总发票数量'] > 0,
            results['作废发票数量'] / results['总发票数量'] * 100,
            0
        )
    else:
        # 如果没有作废发票数据，基于主营业务收入估算
        # 假设每100万收入对应100张发票，作废率2%
        results['总发票数量'] = (results['主营业务收入'] / 10000).astype(int)
        results['作废发票数量'] = (results['总发票数量'] * 0.02).astype(int)
        results['发票作废率'] = 2.0
        print("警告：未找到发票数据，使用估算作废率2%")
    
    # 4. 计算信誉评级
    # 基于毛利率、作废率等指标综合评定
    def calculate_credit_rating(row):
        score = 0
        
        # 毛利率评分（40%权重）
        if row['毛利率'] >= 25:
            score += 40
        elif row['毛利率'] >= 15:
            score += 30
        elif row['毛利率'] >= 5:
            score += 20
        else:
            score += 10
        
        # 发票作废率评分（30%权重）
        if row['发票作废率'] <= 1:
            score += 30
        elif row['发票作废率'] <= 3:
            score += 25
        elif row['发票作废率'] <= 5:
            score += 15
        else:
            score += 5
        
        # 主营业务收入规模评分（30%权重）
        revenue = row['主营业务收入']
        if revenue >= 1000000:  # 100万以上
            score += 30
        elif revenue >= 500000:  # 50万以上
            score += 25
        elif revenue >= 100000:  # 10万以上
            score += 15
        else:
            score += 5
        
        # 评级映射
        if score >= 85:
            return 'AAA'
        elif score >= 75:
            return 'AA'
        elif score >= 65:
            return 'A'
        elif score >= 55:
            return 'BBB'
        elif score >= 45:
            return 'BB'
        elif score >= 35:
            return 'B'
        else:
            return 'C'
    
    results['信誉评级'] = results.apply(calculate_credit_rating, axis=1)
    
    # 格式化数值
    results['主营业务收入'] = results['主营业务收入'].round(2)
    results['主营业务成本'] = results['主营业务成本'].round(2)
    results['毛利率'] = results['毛利率'].round(2)
    results['发票作废率'] = results['发票作废率'].round(2)
    
    return results

def create_styled_excel(results, filename='附件1_企业财务指标分析结果.xlsx'):
    """创建带样式的Excel文件"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "企业财务指标分析"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入标题
    ws.merge_cells('A1:G1')
    ws['A1'] = '附件1企业财务指标分析结果'
    ws['A1'].font = Font(name='微软雅黑', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # 写入表头
    headers = ['企业编号', '主营业务收入(元)', '主营业务成本(元)', '毛利率(%)', '发票作废率(%)', '信誉评级', '综合评分']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for idx, row in results.iterrows():
        # 计算综合评分
        score = 0
        if row['毛利率'] >= 25: score += 40
        elif row['毛利率'] >= 15: score += 30
        elif row['毛利率'] >= 5: score += 20
        else: score += 10
        
        if row['发票作废率'] <= 1: score += 30
        elif row['发票作废率'] <= 3: score += 25
        elif row['发票作废率'] <= 5: score += 15
        else: score += 5
        
        if row['主营业务收入'] >= 1000000: score += 30
        elif row['主营业务收入'] >= 500000: score += 25
        elif row['主营业务收入'] >= 100000: score += 15
        else: score += 5
        
        data_row = [
            row['企业编号'],
            f"{row['主营业务收入']:,.2f}",
            f"{row['主营业务成本']:,.2f}",
            f"{row['毛利率']:.2f}%",
            f"{row['发票作废率']:.2f}%",
            row['信誉评级'],
            score
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws.cell(row=idx+3, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 根据信誉评级设置颜色
            if col == 6:  # 信誉评级列
                if value in ['AAA', 'AA']:
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif value in ['A', 'BBB']:
                    cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                elif value in ['BB', 'B']:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    # 调整列宽
    column_widths = [12, 18, 18, 12, 12, 12, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 添加统计信息
    start_row = len(results) + 5
    ws.merge_cells(f'A{start_row}:G{start_row}')
    ws[f'A{start_row}'] = '统计信息'
    ws[f'A{start_row}'].font = Font(name='微软雅黑', size=14, bold=True)
    ws[f'A{start_row}'].alignment = Alignment(horizontal='center')
    
    stats = [
        ['指标', '平均值', '最大值', '最小值', '标准差'],
        ['主营业务收入', f"{results['主营业务收入'].mean():,.2f}", 
         f"{results['主营业务收入'].max():,.2f}", f"{results['主营业务收入'].min():,.2f}",
         f"{results['主营业务收入'].std():,.2f}"],
        ['毛利率(%)', f"{results['毛利率'].mean():.2f}", 
         f"{results['毛利率'].max():.2f}", f"{results['毛利率'].min():.2f}",
         f"{results['毛利率'].std():.2f}"],
        ['发票作废率(%)', f"{results['发票作废率'].mean():.2f}", 
         f"{results['发票作废率'].max():.2f}", f"{results['发票作废率'].min():.2f}",
         f"{results['发票作废率'].std():.2f}"]
    ]
    
    for i, stat_row in enumerate(stats):
        for j, value in enumerate(stat_row):
            cell = ws.cell(row=start_row+2+i, column=j+1, value=value)
            if i == 0:  # 表头
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # 保存文件
    wb.save(filename)
    print(f"Excel文件已保存：{filename}")

def print_summary(results):
    """打印分析结果摘要"""
    print("\n" + "="*60)
    print("附件1企业财务指标分析结果摘要")
    print("="*60)
    
    print(f"总企业数量：{len(results)}")
    print(f"平均主营业务收入：{results['主营业务收入'].mean():,.2f} 元")
    print(f"平均毛利率：{results['毛利率'].mean():.2f}%")
    print(f"平均发票作废率：{results['发票作废率'].mean():.2f}%")
    
    print("\n信誉评级分布：")
    rating_counts = results['信誉评级'].value_counts().sort_index()
    for rating, count in rating_counts.items():
        percentage = count / len(results) * 100
        print(f"  {rating}级：{count}家 ({percentage:.1f}%)")
    
    print("\n主营业务收入区间分布：")
    revenue_ranges = [
        (0, 100000, "10万以下"),
        (100000, 500000, "10-50万"),
        (500000, 1000000, "50-100万"),
        (1000000, float('inf'), "100万以上")
    ]
    
    for min_val, max_val, label in revenue_ranges:
        if max_val == float('inf'):
            count = (results['主营业务收入'] >= min_val).sum()
        else:
            count = ((results['主营业务收入'] >= min_val) & 
                    (results['主营业务收入'] < max_val)).sum()
        percentage = count / len(results) * 100
        print(f"  {label}：{count}家 ({percentage:.1f}%)")

def main():
    """主函数"""
    print("开始处理附件1企业财务指标计算...")
    
    # 加载数据
    df = load_data()
    if df is None:
        return
    
    # 计算财务指标
    print("\n正在计算财务指标...")
    results = calculate_business_indicators(df)
    
    # 打印摘要
    print_summary(results)
    
    # 创建Excel文件
    print("\n正在生成Excel报告...")
    create_styled_excel(results)
    
    print("\n处理完成！")
    print("输出文件：附件1_企业财务指标分析结果.xlsx")

if __name__ == "__main__":
    main()
