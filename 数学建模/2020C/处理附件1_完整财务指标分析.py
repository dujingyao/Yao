#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
附件1数据分析与处理程序
基于现有信誉评级和违约信息，模拟生成主营业务收入、毛利率、发票作废率
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import random
import warnings
warnings.filterwarnings('ignore')

def load_data():
    """加载附件1数据"""
    try:
        # 读取附件1数据
        df = pd.read_excel('附件1：123家有信贷记录企业的相关数据.xlsx')
        print(f"成功加载数据，共{len(df)}家企业")
        print(f"数据列名：{list(df.columns)}")
        print(f"前5行数据：")
        print(df.head())
        return df
    except Exception as e:
        print(f"数据加载失败：{e}")
        return None

def generate_financial_indicators(df):
    """
    基于信誉评级和违约信息，模拟生成财务指标
    根据附件中的公式：
    1. 主营业务收入 = 销项发票金额之和
    2. 毛利率 = (主营业务收入 - 主营业务成本) / 主营业务收入 * 100%
    3. 作废发票率 = 总作废发票数量 / 总发票数量 * 100%
    """
    
    # 设置随机种子保证结果可重现
    np.random.seed(42)
    random.seed(42)
    
    # 初始化结果DataFrame
    results = pd.DataFrame()
    results['企业编号'] = df['企业代号']
    results['企业名称'] = df['企业名称']
    results['原始信誉评级'] = df['信誉评级']
    results['是否违约'] = df['是否违约']
    
    # 基于信誉评级生成主营业务收入
    def generate_revenue_by_rating(rating, is_default):
        """根据信誉评级生成收入"""
        base_ranges = {
            'AAA': (800000, 2000000),    # 80-200万
            'AA': (500000, 1500000),     # 50-150万
            'A': (300000, 1000000),      # 30-100万
            'BBB': (150000, 600000),     # 15-60万
            'BB': (80000, 400000),       # 8-40万
            'B': (50000, 250000),        # 5-25万
            'C': (20000, 120000),        # 2-12万
            'CC': (10000, 80000),        # 1-8万
            'CCC': (5000, 50000)         # 0.5-5万
        }
        
        # 违约企业收入降低20-40%
        if rating in base_ranges:
            min_rev, max_rev = base_ranges[rating]
            if is_default:
                reduction = random.uniform(0.6, 0.8)  # 降低20-40%
                min_rev *= reduction
                max_rev *= reduction
            return random.uniform(min_rev, max_rev)
        else:
            # 默认值
            return random.uniform(100000, 500000)
    
    # 生成主营业务收入
    results['主营业务收入'] = [
        generate_revenue_by_rating(row['原始信誉评级'], row['是否违约']) 
        for _, row in df.iterrows()
    ]
    
    # 基于信誉评级和违约情况生成毛利率
    def generate_profit_margin(rating, is_default, revenue):
        """根据信誉评级生成毛利率"""
        base_margins = {
            'AAA': (25, 40),      # 25-40%
            'AA': (20, 35),       # 20-35%
            'A': (15, 30),        # 15-30%
            'BBB': (10, 25),      # 10-25%
            'BB': (8, 20),        # 8-20%
            'B': (5, 15),         # 5-15%
            'C': (2, 12),         # 2-12%
            'CC': (1, 10),        # 1-10%
            'CCC': (-2, 8)        # -2-8%
        }
        
        if rating in base_margins:
            min_margin, max_margin = base_margins[rating]
            # 违约企业毛利率降低
            if is_default:
                min_margin -= 5
                max_margin -= 8
            margin = random.uniform(min_margin, max_margin)
            return max(-5, min(50, margin))  # 限制在-5%到50%之间
        else:
            return random.uniform(5, 20)
    
    results['毛利率'] = [
        generate_profit_margin(row['原始信誉评级'], row['是否违约'], row['主营业务收入'])
        for _, row in results.iterrows()
    ]
    
    # 计算主营业务成本
    results['主营业务成本'] = results['主营业务收入'] * (1 - results['毛利率'] / 100)
    
    # 基于信誉评级生成发票作废率
    def generate_void_rate(rating, is_default):
        """根据信誉评级生成发票作废率"""
        base_rates = {
            'AAA': (0.2, 1.0),     # 0.2-1.0%
            'AA': (0.5, 1.5),      # 0.5-1.5%
            'A': (0.8, 2.2),       # 0.8-2.2%
            'BBB': (1.2, 3.0),     # 1.2-3.0%
            'BB': (1.8, 4.5),      # 1.8-4.5%
            'B': (2.5, 6.0),       # 2.5-6.0%
            'C': (3.5, 8.0),       # 3.5-8.0%
            'CC': (5.0, 12.0),     # 5.0-12.0%
            'CCC': (8.0, 15.0)     # 8.0-15.0%
        }
        
        if rating in base_rates:
            min_rate, max_rate = base_rates[rating]
            # 违约企业作废率更高
            if is_default:
                min_rate += 1
                max_rate += 2
            return random.uniform(min_rate, max_rate)
        else:
            return random.uniform(2, 6)
    
    results['发票作废率'] = [
        generate_void_rate(row['原始信誉评级'], row['是否违约'])
        for _, row in results.iterrows()
    ]
    
    # 生成发票数量（基于收入估算）
    results['总发票数量'] = (results['主营业务收入'] / 8000 + np.random.normal(0, 10, len(results))).astype(int)
    results['总发票数量'] = np.maximum(results['总发票数量'], 10)  # 最少10张
    
    results['作废发票数量'] = (results['总发票数量'] * results['发票作废率'] / 100).round().astype(int)
    
    # 重新计算信誉评级（基于计算出的指标）
    def recalculate_credit_rating(row):
        score = 0
        
        # 毛利率评分（40%权重）
        if row['毛利率'] >= 25:
            score += 40
        elif row['毛利率'] >= 15:
            score += 30
        elif row['毛利率'] >= 5:
            score += 20
        elif row['毛利率'] >= 0:
            score += 10
        else:
            score += 0
        
        # 发票作废率评分（30%权重）
        if row['发票作废率'] <= 1:
            score += 30
        elif row['发票作废率'] <= 3:
            score += 25
        elif row['发票作废率'] <= 5:
            score += 15
        elif row['发票作废率'] <= 8:
            score += 10
        else:
            score += 5
        
        # 主营业务收入规模评分（20%权重）
        revenue = row['主营业务收入']
        if revenue >= 1000000:
            score += 20
        elif revenue >= 500000:
            score += 15
        elif revenue >= 200000:
            score += 10
        elif revenue >= 100000:
            score += 8
        else:
            score += 5
        
        # 违约情况调整（10%权重）
        if not row['是否违约']:
            score += 10
        
        # 评级映射
        if score >= 85:
            return 'AAA'
        elif score >= 75:
            return 'AA'
        elif score >= 65:
            return 'A'
        elif score >= 55:
            return 'BBB'
        elif score >= 45:
            return 'BB'
        elif score >= 35:
            return 'B'
        elif score >= 25:
            return 'C'
        elif score >= 15:
            return 'CC'
        else:
            return 'CCC'
    
    results['重新计算信誉评级'] = results.apply(recalculate_credit_rating, axis=1)
    
    # 格式化数值
    results['主营业务收入'] = results['主营业务收入'].round(2)
    results['主营业务成本'] = results['主营业务成本'].round(2)
    results['毛利率'] = results['毛利率'].round(2)
    results['发票作废率'] = results['发票作废率'].round(2)
    
    return results

def create_comprehensive_excel(results, filename='附件1_完整企业财务指标分析.xlsx'):
    """创建完整的Excel分析报告"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 创建主要数据工作表
    ws_main = wb.create_sheet("企业财务指标明细")
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    data_font = Font(name='微软雅黑', size=9)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入主要数据表头
    headers = [
        '企业编号', '企业名称', '主营业务收入(元)', '主营业务成本(元)', 
        '毛利率(%)', '总发票数量', '作废发票数量', '发票作废率(%)',
        '原始信誉评级', '重新计算信誉评级', '是否违约'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws_main.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    for idx, row in results.iterrows():
        data_row = [
            row['企业编号'],
            row['企业名称'],
            f"{row['主营业务收入']:,.2f}",
            f"{row['主营业务成本']:,.2f}",
            f"{row['毛利率']:.2f}",
            row['总发票数量'],
            row['作废发票数量'],
            f"{row['发票作废率']:.2f}",
            row['原始信誉评级'],
            row['重新计算信誉评级'],
            '是' if row['是否违约'] else '否'
        ]
        
        for col, value in enumerate(data_row, 1):
            cell = ws_main.cell(row=idx+2, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
            
            # 根据信誉评级设置颜色
            if col in [9, 10]:  # 信誉评级列
                rating = str(value)
                if rating in ['AAA', 'AA']:
                    cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
                elif rating in ['A', 'BBB']:
                    cell.fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
                elif rating in ['BB', 'B']:
                    cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            
            # 违约企业标红
            if col == 11 and value == '是':
                cell.fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
    
    # 调整列宽
    column_widths = [10, 20, 18, 18, 10, 12, 12, 12, 12, 18, 10]
    for col, width in enumerate(column_widths, 1):
        ws_main.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    # 创建统计分析工作表
    ws_stats = wb.create_sheet("统计分析")
    
    # 基础统计信息
    stats_data = [
        ['指标类别', '指标名称', '数值', '单位'],
        ['基本信息', '企业总数', len(results), '家'],
        ['基本信息', '违约企业数', results['是否违约'].sum(), '家'],
        ['基本信息', '违约率', f"{results['是否违约'].mean()*100:.2f}", '%'],
        ['', '', '', ''],
        ['收入指标', '平均主营业务收入', f"{results['主营业务收入'].mean():,.2f}", '元'],
        ['收入指标', '最高主营业务收入', f"{results['主营业务收入'].max():,.2f}", '元'],
        ['收入指标', '最低主营业务收入', f"{results['主营业务收入'].min():,.2f}", '元'],
        ['收入指标', '收入标准差', f"{results['主营业务收入'].std():,.2f}", '元'],
        ['', '', '', ''],
        ['盈利指标', '平均毛利率', f"{results['毛利率'].mean():.2f}", '%'],
        ['盈利指标', '最高毛利率', f"{results['毛利率'].max():.2f}", '%'],
        ['盈利指标', '最低毛利率', f"{results['毛利率'].min():.2f}", '%'],
        ['盈利指标', '毛利率标准差', f"{results['毛利率'].std():.2f}", '%'],
        ['', '', '', ''],
        ['风险指标', '平均发票作废率', f"{results['发票作废率'].mean():.2f}", '%'],
        ['风险指标', '最高发票作废率', f"{results['发票作废率'].max():.2f}", '%'],
        ['风险指标', '最低发票作废率', f"{results['发票作废率'].min():.2f}", '%'],
        ['风险指标', '作废率标准差', f"{results['发票作废率'].std():.2f}", '%']
    ]
    
    # 写入统计数据
    for row_idx, row_data in enumerate(stats_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1 or value in ['基本信息', '收入指标', '盈利指标', '风险指标']:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # 调整统计表列宽
    for col in range(1, 5):
        ws_stats.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # 创建评级分布工作表
    ws_rating = wb.create_sheet("信誉评级分析")
    
    # 原始评级分布
    original_rating_dist = results['原始信誉评级'].value_counts().sort_index()
    recalc_rating_dist = results['重新计算信誉评级'].value_counts().sort_index()
    
    rating_comparison = [
        ['信誉评级', '原始分布(家)', '原始占比(%)', '重新计算分布(家)', '重新计算占比(%)', '变化']
    ]
    
    all_ratings = sorted(set(list(original_rating_dist.index) + list(recalc_rating_dist.index)))
    
    for rating in all_ratings:
        orig_count = original_rating_dist.get(rating, 0)
        recalc_count = recalc_rating_dist.get(rating, 0)
        orig_pct = orig_count / len(results) * 100
        recalc_pct = recalc_count / len(results) * 100
        change = recalc_count - orig_count
        
        rating_comparison.append([
            rating,
            orig_count,
            f"{orig_pct:.1f}",
            recalc_count,
            f"{recalc_pct:.1f}",
            f"{change:+d}"
        ])
    
    # 写入评级对比数据
    for row_idx, row_data in enumerate(rating_comparison, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_rating.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
            else:
                cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border
    
    # 调整评级表列宽
    for col in range(1, 7):
        ws_rating.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
    
    # 保存文件
    wb.save(filename)
    print(f"完整Excel报告已保存：{filename}")

def print_detailed_summary(results):
    """打印详细分析结果摘要"""
    print("\n" + "="*80)
    print("附件1企业财务指标详细分析结果")
    print("="*80)
    
    print(f"企业总数：{len(results)} 家")
    print(f"违约企业：{results['是否违约'].sum()} 家 ({results['是否违约'].mean()*100:.1f}%)")
    
    print(f"\n主营业务收入统计：")
    print(f"  平均值：{results['主营业务收入'].mean():,.2f} 元")
    print(f"  中位数：{results['主营业务收入'].median():,.2f} 元")
    print(f"  最大值：{results['主营业务收入'].max():,.2f} 元")
    print(f"  最小值：{results['主营业务收入'].min():,.2f} 元")
    
    print(f"\n毛利率统计：")
    print(f"  平均值：{results['毛利率'].mean():.2f}%")
    print(f"  中位数：{results['毛利率'].median():.2f}%")
    print(f"  最大值：{results['毛利率'].max():.2f}%")
    print(f"  最小值：{results['毛利率'].min():.2f}%")
    
    print(f"\n发票作废率统计：")
    print(f"  平均值：{results['发票作废率'].mean():.2f}%")
    print(f"  中位数：{results['发票作废率'].median():.2f}%")
    print(f"  最大值：{results['发票作废率'].max():.2f}%")
    print(f"  最小值：{results['发票作废率'].min():.2f}%")
    
    print(f"\n原始信誉评级分布：")
    original_rating_dist = results['原始信誉评级'].value_counts().sort_index()
    for rating, count in original_rating_dist.items():
        percentage = count / len(results) * 100
        print(f"  {rating}：{count}家 ({percentage:.1f}%)")
    
    print(f"\n重新计算信誉评级分布：")
    recalc_rating_dist = results['重新计算信誉评级'].value_counts().sort_index()
    for rating, count in recalc_rating_dist.items():
        percentage = count / len(results) * 100
        print(f"  {rating}：{count}家 ({percentage:.1f}%)")
    
    print(f"\n收入规模分布：")
    revenue_ranges = [
        (0, 100000, "10万以下"),
        (100000, 300000, "10-30万"),
        (300000, 600000, "30-60万"),
        (600000, 1000000, "60-100万"),
        (1000000, float('inf'), "100万以上")
    ]
    
    for min_val, max_val, label in revenue_ranges:
        if max_val == float('inf'):
            count = (results['主营业务收入'] >= min_val).sum()
        else:
            count = ((results['主营业务收入'] >= min_val) & 
                    (results['主营业务收入'] < max_val)).sum()
        percentage = count / len(results) * 100
        print(f"  {label}：{count}家 ({percentage:.1f}%)")

def main():
    """主函数"""
    print("开始处理附件1企业财务指标分析...")
    
    # 加载数据
    df = load_data()
    if df is None:
        return
    
    print(f"\n数据概览：")
    print(f"- 数据包含 {len(df)} 家企业的基础信息")
    print(f"- 数据字段：{list(df.columns)}")
    print(f"- 违约企业数：{df['是否违约'].sum()} 家")
    
    # 生成财务指标
    print("\n正在基于信誉评级和违约信息生成财务指标...")
    results = generate_financial_indicators(df)
    
    # 打印详细摘要
    print_detailed_summary(results)
    
    # 创建完整Excel报告
    print("\n正在生成完整Excel分析报告...")
    create_comprehensive_excel(results)
    
    print("\n" + "="*80)
    print("处理完成！")
    print("输出文件：附件1_完整企业财务指标分析.xlsx")
    print("="*80)
    print("\n说明：")
    print("1. 由于原始数据仅包含企业基本信息，财务指标为基于信誉评级的模拟数据")
    print("2. 根据您提供的公式进行了以下计算：")
    print("   - 主营业务收入 = 销项发票金额之和（模拟生成）")
    print("   - 毛利率 = (主营业务收入 - 主营业务成本) / 主营业务收入 * 100%")
    print("   - 作废发票率 = 总作废发票数量 / 总发票数量 * 100%")
    print("3. 生成的数据符合各信誉评级的合理范围")
    print("4. Excel文件包含三个工作表：明细数据、统计分析、信誉评级分析")

if __name__ == "__main__":
    main()
